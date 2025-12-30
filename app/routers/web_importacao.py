from __future__ import annotations

import io
from typing import Any

from fastapi import APIRouter, Request, UploadFile, File
from fastapi.responses import HTMLResponse
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from app.database import get_db
from app.models import Empresa, FuncionarioAutorizado

router = APIRouter(tags=["Web"])


def _norm(s: Any) -> str:
    return str(s or "").strip()


def _cpf_digits(cpf: str) -> str:
    # remove tudo que não é número
    return "".join(ch for ch in cpf if ch.isdigit())


@router.get("/painel/importacao", response_class=HTMLResponse)
def importacao_get(request: Request):
    templates = request.app.state.templates
    return templates.TemplateResponse(
        "importacao.html",
        {"request": request, "title": "Importação"},
    )


@router.post("/painel/importacao", response_class=HTMLResponse)
async def importacao_post(
    request: Request,
    file: UploadFile = File(...),
    db: Session = next(get_db()),  # evita Depends em função async com yield
):
    """
    Planilha XLSX com colunas (linha 1):
      empresa_nome | empresa_cnpj | funcionario_nome | funcionario_cpf | funcionario_email | ativo

    - Upsert Empresa por nome (unique).
    - Upsert FuncionarioAutorizado por (empresa_id + cpf).
    """
    templates = request.app.state.templates

    if not file.filename.lower().endswith(".xlsx"):
        return templates.TemplateResponse(
            "importacao.html",
            {
                "request": request,
                "title": "Importação",
                "error": "Envie um arquivo .xlsx (Excel).",
            },
            status_code=400,
        )

    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active

    # Cabeçalhos
    headers = [(_norm(c).lower()) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    required = ["empresa_nome", "funcionario_nome", "funcionario_cpf"]

    missing = [h for h in required if h not in headers]
    if missing:
        return templates.TemplateResponse(
            "importacao.html",
            {
                "request": request,
                "title": "Importação",
                "error": f"Colunas obrigatórias ausentes: {', '.join(missing)}",
            },
            status_code=400,
        )

    idx = {h: headers.index(h) for h in headers}

    created_empresas = 0
    updated_empresas = 0
    created_funcs = 0
    updated_funcs = 0
    errors: list[str] = []

    for row_i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        empresa_nome = _norm(row[idx["empresa_nome"]]) if "empresa_nome" in idx else ""
        empresa_cnpj = _norm(row[idx.get("empresa_cnpj", -1)]) if "empresa_cnpj" in idx and idx["empresa_cnpj"] >= 0 else ""
        func_nome = _norm(row[idx["funcionario_nome"]]) if "funcionario_nome" in idx else ""
        func_cpf = _cpf_digits(_norm(row[idx["funcionario_cpf"]])) if "funcionario_cpf" in idx else ""
        func_email = _norm(row[idx.get("funcionario_email", -1)]) if "funcionario_email" in idx and idx["funcionario_email"] >= 0 else ""
        ativo_raw = _norm(row[idx.get("ativo", -1)]) if "ativo" in idx and idx["ativo"] >= 0 else "true"
        ativo = ativo_raw.lower() not in ("0", "false", "nao", "não", "n", "inativo")

        if not empresa_nome or not func_nome or not func_cpf:
            errors.append(f"Linha {row_i}: empresa_nome/funcionario_nome/funcionario_cpf são obrigatórios.")
            continue

        # 1) Empresa (upsert por nome)
        empresa = db.query(Empresa).filter(Empresa.nome == empresa_nome).first()
        if not empresa:
            empresa = Empresa(nome=empresa_nome, cnpj=empresa_cnpj or None, ativo=True)
            db.add(empresa)
            db.flush()
            created_empresas += 1
        else:
            changed = False
            if empresa_cnpj and empresa.cnpj != empresa_cnpj:
                empresa.cnpj = empresa_cnpj
                changed = True
            if empresa.ativo is False:
                empresa.ativo = True
                changed = True
            if changed:
                updated_empresas += 1

        # 2) Funcionário autorizado (upsert por empresa_id + cpf)
        func = (
            db.query(FuncionarioAutorizado)
            .filter(
                FuncionarioAutorizado.empresa_id == empresa.id,
                FuncionarioAutorizado.cpf == func_cpf,
            )
            .first()
        )

        if not func:
            func = FuncionarioAutorizado(
                nome=func_nome,
                cpf=func_cpf,
                email=func_email or None,
                ativo=ativo,
                empresa_id=empresa.id,
            )
            db.add(func)
            created_funcs += 1
        else:
            changed = False
            if func.nome != func_nome:
                func.nome = func_nome
                changed = True
            if func_email and func.email != func_email:
                func.email = func_email
                changed = True
            if func.ativo != ativo:
                func.ativo = ativo
                changed = True
            if changed:
                updated_funcs += 1

    db.commit()

    return templates.TemplateResponse(
        "importacao.html",
        {
            "request": request,
            "title": "Importação",
            "success": True,
            "stats": {
                "empresas_criadas": created_empresas,
                "empresas_atualizadas": updated_empresas,
                "funcionarios_criados": created_funcs,
                "funcionarios_atualizados": updated_funcs,
                "erros": errors,
            },
        },
    )
